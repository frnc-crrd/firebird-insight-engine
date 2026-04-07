"""Pipeline principal de auditoria CxC para Microsip.

Orquesta la extraccion de datos desde Firebird, generacion del reporte
operativo, auditoria de anomalias, analisis de cartera, KPIs estrategicos
y exportacion a tres archivos Excel independientes.

Los tres archivos generados son:
    00_auditoria_cxc_TIMESTAMP.xlsx
    01_reporte_cxc_TIMESTAMP.xlsx
    02_analisis_cxc_TIMESTAMP.xlsx

Uso:
    python main.py                    # Pipeline completo
    python main.py --test-connection  # Solo probar conexion a Firebird
    python main.py --skip-audit       # Saltar auditoria de anomalias
    python main.py --skip-analytics   # Saltar analisis de cartera
    python main.py --skip-kpis        # Saltar KPIs estrategicos
"""

from __future__ import annotations

import argparse
import logging
import sys
import tempfile
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config.settings import (
    ANOMALIAS,
    EXCEL_ENGINE,
    EXCEL_NOMBRES,
    FIREBIRD_CONFIG,
    KPI_PERIODO_DIAS,
    OUTPUT_DIR,
    RANGOS_ANTIGUEDAD,
    SHEET_PASSWORDS,
)
from src.analytics import Analytics
from src.auditor import Auditor
from src.db_connector import FirebirdConnector
from src.data_transformer import DataTransformer
from src.kpis import generar_kpis
from src.reporte_cxc import agregar_bandas_grupo, generar_reporte_cxc
from src.reporte_pdf import generar_reporte_pdf

# ======================================================================
# LOGGING
# ======================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("main")

# ======================================================================
# CONSTANTES DE FORMATO Y ESTILO
# ======================================================================

COLUMNAS_MONEDA: set[str] = {
    "CARGOS", "ABONOS", "IMPORTE", "IMPUESTO",
    "SALDO_FACTURA", "SALDO_CLIENTE",
    "IMPORTE_TOTAL", "IMPORTE_PROMEDIO", "IMPORTE_MAX",
    "TOTAL_CARGOS", "TOTAL_ABONOS", "SALDO",
    "MONTO_CARGO", "MONTO_ABONOS", "DISPONIBLE",
    "SALDO_TOTAL", "SALDO_VIGENTE", "SALDO_VENCIDO",
    "LIMITE_CREDITO", "SALDO_PENDIENTE", "FACTURAS_PAGADAS",
    "FACTURAS_VIGENTES", "IMPUESTO_TOTAL", "MONTO_TOTAL",
    "IMPORTE_AJUSTE", "VENDIDO", "PAGADO",
    "TOTAL_CARGOS_CANCELADOS", "TOTAL_ABONOS_CANCELADOS"
}

_COLUMNAS_MONEDA_PREFIJOS: tuple[str, ...] = ("FACTURAS_VENCIDAS", "VIGENTE:", "VENCIDO:")

COLUMNAS_FECHA: set[str] = {
    "FECHA_EMISION", "FECHA_VENCIMIENTO",
    "FECHA_HORA_CREACION", "FECHA_HORA_ULT_MODIF",
    "FECHA_HORA_CANCELACION",
}

COLUMNAS_ENTERO: set[str] = {
    "NUM_DOCUMENTOS", "NUM_REGISTROS", "NUM_CARGOS", "NUM_ABONOS",
    "NUM_FACTURAS", "NUM_VENCIDAS", "DIAS_VENCIDO_MAX",
    "NUM_FACTURAS_PENDIENTES", "NUM_FACTURAS_TOTALES",
    "NUM_FACTURAS_VIGENTES", "NUM_FACTURAS_VENCIDAS"
}

COLUMNAS_PORCENTAJE: set[str] = {"PCT_DEL_TOTAL", "UTILIZACION_PCT", "PCT_ACUMULADO", "PCT_VENCIDO", "VALOR"}

# AGRUPACIONES VISUALES POR COLOR
COLS_COLOR_CARGOS = {"TOTAL_CARGOS", "TOTAL_CARGOS_CANCELADOS", "VENDIDO", "CARGOS"}
COLS_COLOR_ABONOS = {"TOTAL_ABONOS", "TOTAL_ABONOS_CANCELADOS", "PAGADO", "ABONOS", "FACTURAS_PAGADAS"}
COLS_COLOR_SALDOS = {"SALDO_PENDIENTE", "SALDO_VIGENTE", "SALDO_VENCIDO", "SALDO_TOTAL", "SALDO", "DISPONIBLE", "LIMITE_CREDITO", "IMPORTE_AJUSTE"}

PESTANAS_PROTEGIDAS: set[str] = {"registros_totales_cxc"}

_CANCELADO_VALUES: list[Any] = ["S", "SI", "s", "si", 1, True, "1"]

# Tipografía unificada a Cambria con recuperación del efecto Muted para los ceros
_FONT_NAME = "Cambria"
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=11)
_FONT_TOTAL = Font(name=_FONT_NAME, bold=True, size=11)
_FONT_NORMAL = Font(name=_FONT_NAME, size=11)
_FONT_MUTED = Font(name=_FONT_NAME, color="808080", size=11)

_HEADER_FILL: PatternFill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CALC_HEADER_FILL: PatternFill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

_BAND_FILL: PatternFill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_WHITE_FILL: PatternFill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

_FILL_TOTAL = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
_FILL_ZERO  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Rellenos semánticos formales para bloques
FILL_AZUL     = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_VERDE    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_ROJO     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Clasificacion ABC
_FILL_A = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_B = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_C = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

_HEADER_ALIGNMENT: Alignment = Alignment(horizontal="center", vertical="center")
_THIN_BORDER: Border = Border(
    left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"), bottom=Side(style="thin", color="B4C6E7"),
)

COLUMNAS_CALCULADAS_CXC: set[str] = {
    "SALDO_FACTURA", "SALDO_CLIENTE", "DELTA_RECAUDO", "ZSCORE_DELTA_RECAUDO",
    "ATIPICO_DELTA_RECAUDO", "CATEGORIA_RECAUDO", "DELTA_MORA",
    "ZSCORE_DELTA_MORA", "ATIPICO_DELTA_MORA", "CATEGORIA_MORA",
    "ZSCORE_IMPORTE", "ATIPICO_IMPORTE",
}

# ======================================================================
# PREPARACION DE DATOS
# ======================================================================

def _formatear_hora(valor: Any) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    if isinstance(valor, dt_time):
        return valor.strftime("%H:%M:%S")
    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M:%S")
    return str(valor)

def _normalizar_fechas_hora(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["FECHA_EMISION", "FECHA_VENCIMIENTO"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "HORA" in df.columns:
        df["HORA"] = df["HORA"].apply(_formatear_hora)
    return df

def preparar_registros_totales(df: pd.DataFrame) -> pd.DataFrame:
    return agregar_bandas_grupo(_normalizar_fechas_hora(df))

def _filtrar_por_acreditar(df_totales: pd.DataFrame) -> pd.DataFrame:
    if "TIPO_IMPTE" not in df_totales.columns:
        return pd.DataFrame()
    tipo_norm = df_totales["TIPO_IMPTE"].astype(str).str.strip().str.upper()
    mask_tipo_a = tipo_norm == "A"
    
    if "CANCELADO" in df_totales.columns:
        mask_activos = ~df_totales["CANCELADO"].isin(_CANCELADO_VALUES)
        resultado = df_totales[mask_tipo_a & mask_activos].copy()
    else:
        resultado = df_totales[mask_tipo_a].copy()
        
    if "_BAND_GROUP" in resultado.columns:
        resultado = resultado.drop(columns=["_BAND_GROUP"])
    resultado = agregar_bandas_grupo(resultado)
    logger.info("Registros por acreditar: %d filas.", len(resultado))
    return resultado

def _filtrar_cancelados(df_totales: pd.DataFrame) -> pd.DataFrame:
    if "CANCELADO" not in df_totales.columns:
        return pd.DataFrame()
    resultado = df_totales[df_totales["CANCELADO"].isin(_CANCELADO_VALUES)].copy()
    
    if "_BAND_GROUP" in resultado.columns:
        resultado = resultado.drop(columns=["_BAND_GROUP"])
    resultado = agregar_bandas_grupo(resultado)
    logger.info("Registros cancelados: %d filas.", len(resultado))
    return resultado

# ======================================================================
# FORMATO EXCEL — FUNCIONES INTERNAS
# ======================================================================

def _aplicar_formato_encabezado(ws: Any, n_cols: int, calc_cols: set[str] | None = None) -> None:
    calc_upper: set[str] = {c.upper() for c in calc_cols} if calc_cols else set()
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        nombre = str(cell.value).upper() if cell.value else ""
        cell.font = _HEADER_FONT
        cell.fill = _CALC_HEADER_FILL if nombre in calc_upper else _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

def _aplicar_bordes_y_fuente(ws: Any, n_filas: int, n_cols: int) -> None:
    for row_idx in range(2, n_filas + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.font = _FONT_NORMAL

def _aplicar_formatos_columna(ws: Any, columnas: list[str], n_filas: int, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(columnas, start=1):
        col_upper = str(col_name).upper()
        es_moneda = (
            col_upper in COLUMNAS_MONEDA
            or any(col_upper.startswith(p) for p in _COLUMNAS_MONEDA_PREFIJOS)
        )
        if es_moneda:
            for row_idx in range(2, n_filas + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
        elif col_upper in COLUMNAS_ENTERO:
            for row_idx in range(2, n_filas + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        elif col_upper in COLUMNAS_FECHA:
            for row_idx in range(2, n_filas + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = "DD/MM/YYYY"
        elif col_upper in COLUMNAS_PORCENTAJE:
            for row_idx in range(2, n_filas + 2):
                aplicar = True
                if col_upper == "VALOR" and "UNIDAD" in df.columns:
                    unidad = df.iloc[row_idx - 2].get("UNIDAD", "")
                    if str(unidad).strip() != "%":
                        aplicar = False
                
                if aplicar:
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

def _aplicar_estilos_semanticos(ws: Any, df: pd.DataFrame, columnas: list[str]) -> None:
    n_cols = len(columnas)
    has_clasif = "CLASIFICACION" in df.columns

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        row_dict = dict(zip(columnas, row))
        
        is_total = str(row_dict.get(columnas[0], "")).strip().upper() == "TOTAL"
        
        is_zero = False
        for saldo_col in ["SALDO_PENDIENTE", "SALDO_TOTAL", "SALDO", "IMPORTE_AJUSTE"]:
            if saldo_col in row_dict:
                val = row_dict[saldo_col]
                if val is not None and str(val).strip() != "":
                    try:
                        if float(val) == 0.0: is_zero = True
                    except: pass
                    
        clasif = str(row_dict.get("CLASIFICACION", ""))

        for c_idx, col in enumerate(columnas, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            
            # 1. Aplicar color semántico base por columna
            if col in ["TOTAL_CARGOS", "TOTAL_CARGOS_CANCELADOS", "LIMITE_CREDITO"]:
                cell.fill = FILL_AZUL
            elif col in ["TOTAL_ABONOS", "TOTAL_ABONOS_CANCELADOS", "SALDO_VIGENTE", "DISPONIBLE", "FACTURAS_PAGADAS"]:
                cell.fill = FILL_VERDE
            elif col in ["SALDO_PENDIENTE", "SALDO_TOTAL", "SALDO", "IMPORTE_AJUSTE"]:
                cell.fill = FILL_AMARILLO
            elif col in ["SALDO_VENCIDO", "DIAS_VENCIDO_MAX", "PCT_VENCIDO"]:
                cell.fill = FILL_ROJO
            else:
                # Bandas alternas base
                if r_idx % 2 == 0: cell.fill = _BAND_FILL
                else: cell.fill = _WHITE_FILL

            # 2. Overrides (Fuerzan sobre la columna)
            if clasif == "A": cell.fill = FILL_VERDE
            elif clasif == "B": cell.fill = FILL_AMARILLO
            elif clasif == "C": cell.fill = FILL_ROJO

            # 3. Muted Rows (Cuentas Saldadas): recuperando el font atenuado gris
            if is_zero:
                cell.fill = _FILL_ZERO
                cell.font = _FONT_MUTED

            # 4. Total row overrides everything
            if is_total:
                cell.fill = _FILL_TOTAL
                cell.font = _FONT_TOTAL

def _aplicar_bandas_alternas(ws: Any, band_data: Any, n_cols: int) -> None:
    for i, band_value in enumerate(band_data):
        row_idx = i + 2
        fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid") if int(band_value) == 0 else _WHITE_FILL
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

def _autoajustar_ancho_columnas(ws: Any) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        header_val = col_cells[0].value
        
        if header_val:
            max_length = len(str(header_val))
            
        for cell in col_cells[1:]:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            fmt = cell.number_format or ""
            val = cell.value
            if "DD/MM/YYYY" in fmt or "YYYY" in fmt:
                cell_len = 10
            elif "#,##0" in fmt or "0.00" in fmt:
                try:
                    cell_len = len(f"{float(val):,.2f}")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            elif "%" in fmt:
                try:
                    cell_len = len(f"{float(val)*100:.2f}%")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            else:
                cell_len = len(str(val))
            if cell_len > max_length:
                max_length = cell_len
                
        # Centrado horizontal y vertical para Interpretacion y Motivo
        if str(header_val).upper() in ["INTERPRETACION", "MOTIVO"]:
            ws.column_dimensions[col_letter].width = 60
            for cell in col_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        else:
            # Ajuste mas holgado (x1.3) para tipografía Cambria, para evitar que corte nombres de columna largos
            adjusted_width = int(max_length * 1.3) + 5
            ws.column_dimensions[col_letter].width = min(max(adjusted_width, 14), 70)

def _extraer_banda(df: pd.DataFrame) -> tuple[pd.DataFrame, Any]:
    if "_BAND_GROUP" in df.columns:
        band_data = df["_BAND_GROUP"].values.copy()
        return df.drop(columns=["_BAND_GROUP"]), band_data
    return df, None

def _escribir_hoja(
    writer: Any, nombre_hoja: str, df: pd.DataFrame, band_data: Any = None,
    protegida: bool = False, password: str = "prac", calc_cols: set[str] | None = None,
) -> None:
    sheet_name = nombre_hoja[:31]
    
    # Se exporta al writer antes de manipular los estilos
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    
    n_filas = len(df)
    n_cols = len(df.columns)
    columnas = [str(c) for c in df.columns]
    
    _aplicar_formato_encabezado(ws, n_cols, calc_cols=calc_cols)
    _aplicar_bordes_y_fuente(ws, n_filas, n_cols)
    _aplicar_formatos_columna(ws, columnas, n_filas, df)
    
    if band_data is not None:
        _aplicar_bandas_alternas(ws, band_data, n_cols)
    else:
        _aplicar_estilos_semanticos(ws, df, columnas)
        
    _autoajustar_ancho_columnas(ws)
    ws.sheet_view.showGridLines = False
    
    if protegida:
        ws.protection.sheet = True
        ws.protection.password = password
        
    logger.info("  Hoja '%s': %d filas%s", sheet_name, n_filas, " (protegida)" if protegida else "")

def _exportar_excel(*args, **kwargs) -> Path:
    """
    Exporta diccionarios de DataFrames a un archivo Excel aplicando formatos estrictos.

    Args:
        *args: Argumentos posicionales.
        **kwargs: Parámetros de configuración incluyendo dataframes, nombre_base, output_dir y timestamp.

    Returns:
        Ruta del archivo generado como objeto Path.
    """
    # 1. Resolucion dinamica del diccionario de datos (DataFrames)
    dataframes = kwargs.get("dataframes") or kwargs.get("datos")
    if dataframes is None:
        for arg in args:
            if isinstance(arg, dict):
                dataframes = arg
                break
    if dataframes is None:
        dataframes = {}

    # 2. Construcción estricta de la ruta absoluta
    nombre_base = kwargs.get("nombre_base", "exportacion")
    timestamp = kwargs.get("timestamp", "")
    output_dir = kwargs.get("output_dir")
    orden_hojas = kwargs.get("orden_hojas", list(dataframes.keys()))
    cols_calc_por_hoja = kwargs.get("cols_calc_por_hoja", {})

    sufijo_tiempo = f"_{timestamp}" if timestamp else ""
    nombre_archivo = f"{nombre_base}{sufijo_tiempo}.xlsx"

    if output_dir:
        filepath = Path(output_dir) / nombre_archivo
        filepath.parent.mkdir(parents=True, exist_ok=True)
    else:
        logger.warning("No se detecto output_dir. Redirigiendo a directorio temporal.")
        filepath = Path(tempfile.gettempdir()) / nombre_archivo

    # 3. Exportacion con salvaguarda de formato e inyección de estilos
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        hojas_escritas = 0
        
        # Procesamiento priorizando el orden solicitado
        for nombre_hoja in orden_hojas:
            if nombre_hoja in dataframes:
                df = dataframes[nombre_hoja]
                if df is not None and not df.empty:
                    df_limpio, band_data = _extraer_banda(df)
                    protegida = nombre_hoja in PESTANAS_PROTEGIDAS
                    calc_cols = cols_calc_por_hoja.get(nombre_hoja)
                    
                    _escribir_hoja(
                        writer, nombre_hoja, df_limpio,
                        band_data=band_data,
                        protegida=protegida,
                        password=SHEET_PASSWORDS.get(nombre_hoja, "prac"),
                        calc_cols=calc_cols
                    )
                    hojas_escritas += 1

        # Agregar hojas restantes no declaradas en el orden
        for nombre_hoja, df in dataframes.items():
            if nombre_hoja not in orden_hojas and df is not None and not df.empty:
                df_limpio, band_data = _extraer_banda(df)
                protegida = nombre_hoja in PESTANAS_PROTEGIDAS
                calc_cols = cols_calc_por_hoja.get(nombre_hoja)
                
                _escribir_hoja(
                    writer, nombre_hoja, df_limpio,
                    band_data=band_data,
                    protegida=protegida,
                    password=SHEET_PASSWORDS.get(nombre_hoja, "prac"),
                    calc_cols=calc_cols
                )
                hojas_escritas += 1
        
        # Contingencia: Prevenir la corrupcion del archivo .xlsx si no hay datos
        if hojas_escritas == 0:
            logger.warning(f"Ninguna hoja activa para {filepath.name}. Inyectando contingencia.")
            df_vacio = pd.DataFrame({"AVISO": ["Ausencia de datos transaccionales en este periodo"]})
            df_vacio.to_excel(writer, sheet_name="Sin Datos", index=False)
            
    return filepath


# ======================================================================
# EXPORTACION — TRES ARCHIVOS EXCEL
# ======================================================================

def exportar_tres_exceles(
    cxc: dict[str, pd.DataFrame],
    auditoria: dict[str, pd.DataFrame],
    analisis: dict[str, pd.DataFrame],
    kpis: dict[str, pd.DataFrame],
    timestamp: str,
    output_dir: Path,
) -> list[Path]:
    archivos: list[Path] = []

    logger.info("Exportando 01_cxc...")
    archivos.append(_exportar_excel(
        dataframes=cxc,
        nombre_base=EXCEL_NOMBRES["cxc"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "movimientos_abiertos_cxc",
            "movimientos_cerrados_cxc",
            "movimientos_totales_cxc",
            "registros_por_acreditar_cxc",
            "registros_cancelados_cxc",
            "registros_totales_cxc",
        ],
        cols_calc_por_hoja={"movimientos_totales_cxc": COLUMNAS_CALCULADAS_CXC},
    ))

    analisis_compilado = {**analisis}
    hojas_kpis_a_fusionar = [
        "kpis_resumen_mxn", "kpis_resumen_usd", 
        "kpis_concentracion_mxn", "kpis_concentracion_usd",
        "kpis_limite_credito_mxn", "kpis_limite_credito_usd",
        "kpis_morosidad_cliente_mxn", "kpis_morosidad_cliente_usd"
    ]
    for hoja_kpi in hojas_kpis_a_fusionar:
        if hoja_kpi in kpis:
            analisis_compilado[hoja_kpi] = kpis.pop(hoja_kpi)

    logger.info("Exportando 02_analisis...")
    archivos.append(_exportar_excel(
        dataframes=analisis_compilado,
        nombre_base=EXCEL_NOMBRES["analisis"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "cartera_vencida_vs_vigente_mxn",
            "cartera_vencida_vs_vigente_usd",
            "antiguedad_cartera_mxn",
            "antiguedad_cartera_usd",
            "antiguedad_por_cliente_mxn",
            "antiguedad_por_cliente_usd",
            "resumen_concepto_cxc_mxn",
            "resumen_concepto_cxc_usd",
            "resumen_cancelados_cxc_mxn",
            "resumen_cancelados_cxc_usd",
            "resumen_ajustes_cxc_mxn",
            "resumen_ajustes_cxc_usd",
            "kpis_resumen_mxn",
            "kpis_resumen_usd",
            "kpis_concentracion_mxn",
            "kpis_concentracion_usd",
            "kpis_limite_credito_mxn",
            "kpis_limite_credito_usd",
            "kpis_morosidad_cliente_mxn",
            "kpis_morosidad_cliente_usd",
        ],
    ))

    logger.info("Exportando 00_auditoria...")
    archivos.append(_exportar_excel(
        dataframes=auditoria,
        nombre_base=EXCEL_NOMBRES["auditoria"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "calidad_datos",
            "importes_atipicos",
            "recaudos_atipicos",
            "moras_atipicas",
            "sin_tipo_cliente",
            "sin_vendedor",
        ],
    ))

    return archivos

# ======================================================================
# PIPELINE
# ======================================================================

def run_pipeline(skip_audit: bool = False, skip_analytics: bool = False, skip_kpis: bool = False) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info("=" * 60)
    logger.info("PASO 1: Extraccion y transformacion de datos en memoria")
    logger.info("=" * 60)

    connector = FirebirdConnector(FIREBIRD_CONFIG)
    try:
        transformer = DataTransformer(connector)
        df = transformer.get_master_cxc_data()
    except Exception as exc:
        logger.error("Error al procesar los datos transaccionales: %s", exc)
        return 1

    logger.info("Datos unificados en memoria: %d filas x %d columnas", *df.shape)

    logger.info("=" * 60)
    logger.info("PASO 2: Reporte operativo CxC")
    logger.info("=" * 60)

    resultado_reporte = generar_reporte_cxc(df)
    registros_totales = preparar_registros_totales(df)
    registros_por_acreditar = _filtrar_por_acreditar(registros_totales)
    registros_cancelados = _filtrar_cancelados(registros_totales)

    cxc: dict[str, pd.DataFrame] = {
        "movimientos_abiertos_cxc":    resultado_reporte.get("movimientos_abiertos_cxc", pd.DataFrame()),
        "movimientos_cerrados_cxc":    resultado_reporte.get("movimientos_cerrados_cxc", pd.DataFrame()),
        "movimientos_totales_cxc":     resultado_reporte.get("movimientos_totales_cxc", pd.DataFrame()),
        "registros_por_acreditar_cxc": registros_por_acreditar,
        "registros_cancelados_cxc":    registros_cancelados,
        "registros_totales_cxc":       registros_totales,
    }

    auditoria: dict[str, pd.DataFrame] = {}
    if not skip_audit:
        logger.info("=" * 60)
        logger.info("PASO 3: Auditoria y deteccion de anomalias")
        logger.info("=" * 60)
        reporte_cxc_df = resultado_reporte.get("reporte_cxc", pd.DataFrame())
        auditor = Auditor(ANOMALIAS)
        audit_result = auditor.run_audit(df, df_reporte=reporte_cxc_df)
        auditoria = {
            "calidad_datos":     audit_result.calidad_datos,
            "importes_atipicos": audit_result.importes_atipicos,
            "recaudos_atipicos": audit_result.recaudos_atipicos,
            "moras_atipicas":    audit_result.moras_atipicas,
            "sin_tipo_cliente":  audit_result.sin_tipo_cliente,
            "sin_vendedor":      audit_result.sin_vendedor,
        }

    analisis: dict[str, pd.DataFrame] = {}
    if not skip_analytics:
        logger.info("=" * 60)
        logger.info("PASO 4: Analisis de cartera")
        logger.info("=" * 60)
        vistas_analytics = {
            "movimientos_abiertos_cxc":    cxc.get("movimientos_abiertos_cxc", pd.DataFrame()),
            "movimientos_totales_cxc":     cxc.get("movimientos_totales_cxc", pd.DataFrame()),
            "registros_por_acreditar_cxc": cxc.get("registros_por_acreditar_cxc", pd.DataFrame()),
            "registros_cancelados_cxc":    cxc.get("registros_cancelados_cxc", pd.DataFrame()),
        }
        analytics_engine = Analytics(RANGOS_ANTIGUEDAD)
        analisis = analytics_engine.run_analytics(vistas_analytics)

    kpis: dict[str, pd.DataFrame] = {}
    if not skip_kpis:
        logger.info("=" * 60)
        logger.info("PASO 5: KPIs estrategicos")
        logger.info("=" * 60)
        kpis = generar_kpis(cxc.get("movimientos_totales_cxc", pd.DataFrame()), KPI_PERIODO_DIAS)

    if not skip_analytics and not skip_kpis:
        analisis_pdf = {**analisis, **kpis}
        logger.info("=" * 60)
        logger.info("PASO 4b: Generando PDF de analisis")
        logger.info("=" * 60)
        try:
            ts_legible = datetime.now().strftime("%Y-%m-%d %H:%M")
            pdf_path = OUTPUT_DIR / f"{EXCEL_NOMBRES['pdf']}_{timestamp}.pdf"
            generar_reporte_pdf(analisis_pdf, pdf_path, ts_legible)
        except Exception as exc:
            logger.warning("No se pudo generar el PDF de analisis: %s", exc)

    logger.info("=" * 60)
    logger.info("PASO 6: Exportacion a tres archivos Excel")
    logger.info("=" * 60)
    archivos_generados = exportar_tres_exceles(
        cxc=cxc, auditoria=auditoria, analisis=analisis, kpis=kpis,
        timestamp=timestamp, output_dir=OUTPUT_DIR,
    )

    logger.info("=" * 60)
    logger.info("PIPELINE COMPLETADO EXITOSAMENTE")
    for archivo in archivos_generados:
        logger.info("  %s", archivo.name)
    logger.info("=" * 60)
    return 0

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pipeline de auditoria CxC para Microsip")
    parser.add_argument("--test-connection", action="store_true", help="Solo probar conexion a Firebird.")
    parser.add_argument("--skip-audit", action="store_true", help="Saltar auditoria de anomalias.")
    parser.add_argument("--skip-analytics", action="store_true", help="Saltar analisis de cartera.")
    parser.add_argument("--skip-kpis", action="store_true", help="Saltar KPIs estrategicos.")
    return parser.parse_args()

def main() -> int:
    args = parse_args()
    if args.test_connection:
        connector = FirebirdConnector(FIREBIRD_CONFIG)
        return 0 if connector.test_connection() else 1
    return run_pipeline(skip_audit=args.skip_audit, skip_analytics=args.skip_analytics, skip_kpis=args.skip_kpis)

if __name__ == "__main__":
    sys.exit(main())